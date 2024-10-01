#!/usr/bin/env python
"""
form2list モジュールは、指定されたディレクトリ内の Excel ファイルを処理し、指定された出力ファイルに一覧を生成する CLI を提供します。
"""

import argparse
import os
import sys
from ast import literal_eval
import openpyxl
import yaml
from jinja2 import Template

def parse_arguments():
    """コマンドライン引数をパースして返す"""
    parser = argparse.ArgumentParser(description='ディレクトリを巡回して入力ファイルを探します。')
    parser.add_argument('-c', '--config', type=str, default='spec.yml', help='変換仕様定義YAMLファイル名')
    parser.add_argument('-o', '--output', type=str, default='list.xlsx', help='出力ファイル')
    parser.add_argument('-v', '--verbose', action='store_true', help='詳細な出力を表示')
    parser.add_argument('directory', type=str, help='読み込みフォルダのパス')
    return parser.parse_args()

def find_input_files(directory):
    """指定されたディレクトリ内の Excel ファイルを再帰的に探してリストで返す"""
    input_files = []
    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith('.xlsx'):  # 入力ファイルの拡張子を指定
                input_files.append(os.path.join(root, file))
    return input_files

def verbose_print(verbose, message):
    """-v オプションが指定されている場合にメッセージを表示する"""
    if verbose:
        print(message)

def column_number_to_name(n):
    """列番号をExcelの列名に変換する"""
    name = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        name = chr(65 + remainder) + name
    return name

def process_file(file_path, config, output_wb, row_number, verbose):
    """指定された Excel ファイルを読み込み、指定されたシートに書き出す"""
    verbose_print(verbose, f'Processing file: {file_path}')
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    context = {}
    found = False
    for spec in config['inputFormats']:
        for key, cell in spec['items'].items():
            context[key] = ws[cell].value
        condition_result = spec['condition_template'].render(context)
        verbose_print(verbose, f"Rendered check condition {spec.get('condition', 'True')} on context {context} is {condition_result}")
        if literal_eval(condition_result):
            found = True
            break
    if not found:
        print(f"Fail to load values from: {file_path}", file=sys.stderr)
        return False
    context['dirname'] = os.path.dirname(file_path)
    context['basename'] = os.path.basename(file_path)

    for sheet_config in config['sheets']:
        context['row_number'] = row_number + int(sheet_config.get('rowOffset', 0))
        context['row'] = str(context['row_number'])
        ws = output_wb[sheet_config['name']]
        column_number = 1
        for col in sheet_config['columns']:
            context['column_number'] = column_number + int(config.get('columnOffset', 0))
            context['column'] = column_number_to_name(context['column_number'])
            context['cell'] = f"{context['column']}{context['row']}"
            verbose_print(verbose, f"Writing to {context['cell']} with template {col['value']} on context {context}")
            cell = ws[context['cell']]
            cell.value = col["value_template"].render(context)
            column_number += 1
    return True

def setup_templates(config):
    """設定ファイルのテンプレートを初期化する"""
    for spec in config['inputFormats']:
        condition = spec.get('condition', 'True')
        spec['condition_template'] = Template(condition)
    for sheet_config in config['sheets']:
        for col in sheet_config.get('columns', []):
            col['value_template'] = Template(col['value'])

def main():
    """階層構造のフォルダ内に格納されている申込書の Excel ファイルを巡回して、申請項目の内容を読み出し、指定された Excel ファイル内に一覧を書き出す"""
    args = parse_arguments()
    input_files = find_input_files(args.directory)
    
    if not input_files:
        print('Error: 入力ファイルが見つかりませんでした。', file=sys.stderr)
        return  1
    
    try:
        with open(args.config, 'r', encoding="utf-8") as yaml_file:
            config = yaml.safe_load(yaml_file)
    except FileNotFoundError:
        print(f'Error: 設定ファイル {args.config} が見つかりませんでした。', file=sys.stderr)
        return 1

    try:
        setup_templates(config)
    except KeyError as e:
        print(f'Error: 仕様書 YAML に必須要素が指定されていません: {e}', file=sys.stderr)
        return 1

    try:
        template_wb = openpyxl.load_workbook(config['template'])
    except FileNotFoundError:
        print(f'Error: テンプレートファイル {config['template']} が見つかりませんでした。', file=sys.stderr)
        return 1
    except KeyError as e:
        print(f'Error: 仕様書 YAML にテンプレートファイルが指定されていません: {e}', file=sys.stderr)
        return 1

    row_number = 1
    for file_path in input_files:
        result = process_file(file_path, config, template_wb, row_number, args.verbose)
        row_number += 1
        if not result:
            return 1
    template_wb.save(args.output)
    return 0

if __name__ == '__main__':
    sys.exit(main())
